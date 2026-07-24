"""2단계 문서 자동화 — 검증 완료 데이터를 엑셀 체크시트로 출력. openpyxl 사용."""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app import db

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")
PASS_FILL = PatternFill("solid", fgColor="DCFCE7")
FAIL_FILL = PatternFill("solid", fgColor="FEE2E2")
ALERT_FILL = PatternFill("solid", fgColor="FEE2E2")
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _run_data(run_id):
    run = db.query(
        "SELECT r.*, t.model_name, t.model_rev, t.tester_type, t.unit_no, t.unit_label, t.customer "
        "FROM inspection_run r JOIN tester t ON t.tester_id = r.tester_id WHERE r.run_id = ?",
        (run_id,), one=True,
    )
    if not run:
        return None
    run["check_items"] = db.query("SELECT * FROM check_item WHERE run_id = ? ORDER BY seq", (run_id,))
    run["measurements"] = db.query(
        "SELECT * FROM measurement WHERE run_id = ? AND repeat_index IS NULL ORDER BY item", (run_id,)
    )
    return run


def _kv_row(ws, row, label, value):
    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.cell(row=row, column=2, value=value)
    for col in (1, 2):
        ws.cell(row=row, column=col).border = BORDER


def build_checksheet(run_id):
    """검증 세션 1건에 대한 체크시트 엑셀(bytes)을 생성. run이 없으면 None."""
    run = _run_data(run_id)
    if not run:
        return None, None

    wb = Workbook()
    ws = wb.active
    ws.title = "체크시트"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 16

    ws.merge_cells("A1:D1")
    ws["A1"] = "KNK 검사기 출하검증 체크시트"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    row = 3
    for label, value in [
        ("모델명 / REV", f"{run['model_name']} / {run['model_rev'] or '-'}"),
        ("검사기 종류 / 호기", f"{run['tester_type']} / {run['unit_no'] or '-'}호기"),
        ("고객사", run["customer"] or "-"),
        ("검사자 / 검증 모드", f"{run['inspector']} / {run['verify_mode']}"),
        ("검증일", run["run_date"]),
        ("종합 판정", run["result"]),
    ]:
        _kv_row(ws, row, label, value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="검사 항목 결과 (Check Sheet)").font = Font(bold=True, size=12)
    row += 1
    headers = ["No", "구분", "항목", "판정"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    row += 1
    for it in run["check_items"]:
        ws.cell(row=row, column=1, value=it["seq"]).border = BORDER
        ws.cell(row=row, column=2, value=it["category"]).border = BORDER
        ws.cell(row=row, column=3, value=it["item_name"]).border = BORDER
        rcell = ws.cell(row=row, column=4, value=it["result"])
        rcell.border = BORDER
        rcell.alignment = Alignment(horizontal="center")
        if it["result"] == "PASS":
            rcell.fill = PASS_FILL
        elif it["result"] == "FAIL":
            rcell.fill = FAIL_FILL
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="측정 데이터 (Log 자동 판정)").font = Font(bold=True, size=12)
    row += 1
    headers = ["측정 항목", "실측값", "규격(low~high)", "판정"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    row += 1
    if run["measurements"]:
        for m in run["measurements"]:
            ws.cell(row=row, column=1, value=m["item"]).border = BORDER
            ws.cell(row=row, column=2, value=m["value"]).border = BORDER
            ws.cell(row=row, column=3, value=f"{m['spec_low']} ~ {m['spec_high']}").border = BORDER
            jcell = ws.cell(row=row, column=4, value=m["judge"])
            jcell.border = BORDER
            jcell.alignment = Alignment(horizontal="center")
            if m["judge"] == "알림":
                jcell.fill = ALERT_FILL
            elif m["judge"] == "주의":
                jcell.fill = WARN_FILL
            row += 1
    else:
        ws.cell(row=row, column=1, value="측정 데이터 없음")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="검사자 의견").font = Font(bold=True, size=12)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=4)
    c = ws.cell(row=row, column=1, value=run["inspector_comment"] or "-")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_model = "".join(ch for ch in run["model_name"] if ch not in '\\/:*?"<>|')
    date_part = (run["run_date"] or "")[:10].replace("-", "")
    filename = f"체크시트_{safe_model}_{date_part}.xlsx"
    return buf.read(), filename


# ---------------------------------------------------------------------------
# 주간 업무 보고서 — 기간별 검증 완료 데이터(검사자 의견)를 카테고리별로 묶어 출력
# ---------------------------------------------------------------------------
WEEKLY_CATEGORIES = [
    ("PBA.", ["기능검사기", "방수", "VSWR", "LNA", "PROXIMITY"]),
    ("TSP.", ["TSP"]),
    ("지문,센서 파트.", ["지문"]),
]


def _weekly_runs(start_date, end_date):
    return db.query(
        """
        SELECT r.run_date, r.inspector, r.inspector_comment, r.result,
               t.model_name, t.tester_type, t.customer
        FROM inspection_run r JOIN tester t ON t.tester_id = r.tester_id
        WHERE date(r.run_date) BETWEEN date(?) AND date(?)
          AND r.inspector_comment IS NOT NULL AND trim(r.inspector_comment) != ''
        ORDER BY r.run_date
        """,
        (start_date, end_date),
    )


def _write_section_header(ws, row, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value=title)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    return row + 1


def _write_bullets(ws, row, text, customer, model_name, tester_type):
    prefix = f"{customer + ' ' if customer else ''}{model_name} {tester_type}"
    lines = [l for l in (text or "").splitlines() if l.strip()]
    if not lines:
        return row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value=f" -. {prefix} — {lines[0]}")
    row += 1
    for extra in lines[1:]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"   -> {extra}")
        row += 1
    return row


def build_weekly_report(start_date, end_date):
    """기간 내 검증완료 건의 검사자 의견을 PBA/TSP/지문·센서 파트로 묶어 주간 업무 보고서 생성."""
    runs = _weekly_runs(start_date, end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "주간업무보고"
    ws.column_dimensions["A"].width = 90
    for col in "BCDEF":
        ws.column_dimensions[col].width = 4

    ws.merge_cells("A1:F1")
    ws["A1"] = f"주간 업무 보고 ({start_date} ~ {end_date})"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    inspectors = sorted({r["inspector"] for r in runs if r["inspector"]})
    ws["A3"] = f"작성자 : {', '.join(inspectors) if inspectors else '-'}"

    row = 5
    row = _write_section_header(ws, row, "전 주 업 무 실 적")

    for cat_title, tester_types in WEEKLY_CATEGORIES:
        row += 1
        ws.cell(row=row, column=1, value=cat_title).font = Font(bold=True)
        row += 1
        cat_runs = [r for r in runs if r["tester_type"] in tester_types]
        if not cat_runs:
            ws.cell(row=row, column=1, value=" -. 해당 기간 내용 없음")
            row += 1
            continue
        for r in cat_runs:
            row = _write_bullets(ws, row, r["inspector_comment"], r["customer"], r["model_name"], r["tester_type"])

    row += 1
    row = _write_section_header(ws, row, "특이 사항")
    fail_runs = [r for r in runs if r["result"] == "FAIL"]
    if not fail_runs:
        row += 1
        ws.cell(row=row, column=1, value=" -. 해당 기간 FAIL 건 없음")
    else:
        for r in fail_runs:
            row += 1
            row = _write_bullets(ws, row, r["inspector_comment"], r["customer"], r["model_name"], r["tester_type"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"주간업무보고_{start_date}_{end_date}.xlsx"
    return buf.read(), filename


# ---------------------------------------------------------------------------
# 히스토리 내보내기 — 검색 필터에 걸린 검증 이력을 목록 형태로 출력
# ---------------------------------------------------------------------------
def build_history_export(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "검증이력"

    headers = ["run_id", "검증일", "모델명", "REV", "검사기종류", "호기", "고객사", "검사자", "모드", "판정", "검사자 의견"]
    widths = [8, 18, 20, 10, 14, 6, 14, 12, 10, 10, 50]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    row = 2
    for r in rows:
        values = [
            r["run_id"], r["run_date"], r["model_name"], r["model_rev"] or "-",
            r["tester_type"], r["unit_no"] or "-", r["customer"] or "-",
            r["inspector"] or "-", r["verify_mode"] or "-", r["result"], r["inspector_comment"] or "",
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = BORDER
            if col == 10:
                if v == "PASS":
                    cell.fill = PASS_FILL
                elif v == "FAIL":
                    cell.fill = FAIL_FILL
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime
    filename = f"검증이력_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buf.read(), filename


def build_issues_export(rows):
    """이슈 이력 목록(필터 반영)을 엑셀로 — 구조화 칸(증상/원인/조치/상태/태그) 포함."""
    wb = Workbook()
    ws = wb.active
    ws.title = "이슈이력"

    headers = ["ID", "검증일", "모델명", "고객사", "검사기종류", "호기", "시료", "증상분류",
               "태그", "제목", "증상", "원인", "조치", "상태"]
    widths = [6, 12, 18, 10, 12, 10, 8, 16, 24, 30, 45, 25, 45, 14]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    row = 2
    for r in rows:
        values = [
            r["id"], (r.get("issue_date") or "")[:10], r["model_name"], r.get("customer") or "-",
            r.get("tester_type") or "-", r.get("unit_label") or "-", r.get("sample_rev") or "-",
            r.get("symptom_type") or "-", (r.get("tags") or "").strip(","),
            r.get("title") or "", r.get("symptom") or "", r.get("cause") or "",
            r.get("action") or "", r.get("status") or "-",
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (11, 13)))
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import datetime
    filename = f"이슈이력_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buf.read(), filename
