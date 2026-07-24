# -*- coding: utf-8 -*-
"""프로그램 → 서버(Z:) 출하이슈사항 엑셀 기록.

이슈 등록(이슈관리 화면·퀵 이슈·새 검증) 시 해당 모델의 서버 출하이슈사항
엑셀에 같은 내용을 한 행으로 기록한다. '누락 금지'가 최우선 원칙:

  · 모델 폴더/이슈 파일을 못 찾으면  → 명확한 오류 반환 (조용한 실패 금지)
  · 엑셀이 열려 있으면(잠김)         → 오류 반환, 닫은 뒤 재시도 유도
  · 저장 후 파일을 다시 열어         → 방금 쓴 행이 실제로 있는지 검증
  · 같은 내용이 이미 있으면          → 중복 기록하지 않고 성공(멱등, 재시도 안전)

파일 형식 (서버의 기존 출하이슈사항정리.xlsx 와 동일):
  A1  '출하 이슈사항 관리' / 2행 헤더(A2 출하 날짜 · E2 내역 · AU2 사진 및 추가 자료)
  3행부터: A{r}(~D 병합)=날짜, E{r}(~AT 병합)=내역, AU{r}=사진
  내역 스타일: "1,2호기 MODIFY\n-.증상…\n->조치…"
"""
import datetime
import os
import re

from app import db

# 사진을 셀에 맞춰 넣을 때의 최대 크기(px) — AU열 너비(66자≈470px) 기준
_IMG_MAX_W = 440
_IMG_MAX_H = 330

_LOCK_MSG = ("서버의 출하이슈사항 엑셀('{name}')이 다른 곳에서 열려 있어 기록할 수 없습니다.\n"
             "엑셀을 닫은 뒤 [다시 시도]를 눌러 주세요.")


# ---------------------------------------------------------------- 대상 파일 찾기
def find_issue_file(model, tester_type=None):
    """모델의 서버 출하이슈사항 엑셀 경로를 찾는다.

    반환: {"ok": True, "path": ..., "tester_dir": ...}
       또는 {"ok": False, "error": ..., "not_found": True}
    """
    from app import zserver
    if not zserver.available():
        return {"ok": False, "error": "사내 서버(Z:)에 접근할 수 없습니다.\n"
                "서버 드라이브가 연결된 회사 PC에서 사용하세요.", "not_found": True}

    zserver.clear_cache()          # 방금 만든 폴더/파일도 바로 보이도록
    assets = zserver.model_assets(model, tester_type)
    if assets.get("available") and not assets.get("testers") and tester_type:
        # 검사기 종류 표기가 폴더와 달라 필터에서 다 빠졌으면 종류 무시하고 재시도
        assets = zserver.model_assets(model, None)
    if not assets.get("available") or not assets.get("testers"):
        return {"ok": False, "not_found": True, "error":
                f"서버에서 '{model}' 모델 폴더를 찾지 못했습니다.\n"
                f"· 서버의 고객사/카테고리 아래 모델 폴더 이름을 확인해 주세요.\n"
                f"· 모델명이 서버 폴더명과 다르면 이슈의 모델명을 폴더명에 맞춰 주세요."}

    for t in assets["testers"]:
        for x in t.get("excels", []):
            if "이슈" in x["name"]:
                return {"ok": True, "path": x["path"], "name": x["name"],
                        "tester_dir": t.get("tester_dir", ""), "locked": x.get("locked")}
    first = assets["testers"][0]
    return {"ok": False, "not_found": True, "error":
            f"'{model}' 모델 폴더는 찾았지만 출하검증 폴더에 출하이슈사항 엑셀이 없습니다.\n"
            f"· 위치: {first.get('verify_dir', first.get('path', ''))}\n"
            f"· '출하이슈사항정리.xlsx' 파일을 만들어 두면 자동 기록됩니다."}


# ---------------------------------------------------------------- 내역 조립
def build_content(issue):
    """이슈 레코드 → 서버 엑셀 '내역' 셀 텍스트 (기존 수기 작성 스타일)."""
    lines = []
    unit = (issue["unit_label"] or "").strip()
    if unit:
        lines.append(unit)
    rev = (issue["sample_rev"] or "").strip()
    if rev:
        lines.append(f"-.{rev} 시료로 검토진행")
    for part in re.split(r"[\r\n]+", (issue["symptom"] or "").strip()):
        p = part.strip()
        if p:
            lines.append(p if p.startswith(("-.", "->")) else f"-.{p}")
    cause = (issue["cause"] or "").strip()
    if cause:
        lines.append(f"-.원인: {cause}")
    action = (issue["action"] or "").strip()
    if action:
        for part in re.split(r"[\r\n]+", action):
            p = part.strip()
            if p:
                lines.append(p if p.startswith("->") else f"->{p}")
    status = (issue["status"] or "").strip()
    if status:
        lines.append(f"({status})")
    return "\n".join(lines)


def _norm_text(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


# ---------------------------------------------------------------- 엑셀 기록
def _next_row(ws):
    """다음 빈 데이터 행(3행부터 A·E열 모두 빈 첫 행)."""
    r = 3
    while r <= ws.max_row + 1:
        a = ws.cell(row=r, column=1).value
        e = ws.cell(row=r, column=5).value
        if (a is None or str(a).strip() == "") and (e is None or str(e).strip() == ""):
            return r
        r += 1
    return r


def _find_duplicate(ws, content):
    """같은 내역이 이미 있는 행 번호(멱등 재시도용). 없으면 None."""
    want = _norm_text(content)
    for r in range(3, ws.max_row + 1):
        if _norm_text(ws.cell(row=r, column=5).value) == want:
            return r
    return None


def _ensure_row_merges(ws, row):
    """대상 행에 A~D·E~AT 병합이 없으면 추가(템플릿 범위를 벗어난 행 대비)."""
    have = {str(m) for m in ws.merged_cells.ranges}
    for rng in (f"A{row}:D{row}", f"E{row}:AT{row}"):
        if rng not in have:
            try:
                ws.merge_cells(rng)
            except ValueError:
                pass                        # 일부 겹침 등 — 값 기록에는 지장 없음


def _fit_image(img_path):
    """openpyxl Image + 셀에 맞는 표시 크기(원본 비율 유지)."""
    from openpyxl.drawing.image import Image as XLImage
    img = XLImage(img_path)
    w, h = img.width or _IMG_MAX_W, img.height or _IMG_MAX_H
    scale = min(_IMG_MAX_W / w, _IMG_MAX_H / h, 1.0)
    img.width, img.height = int(w * scale), int(h * scale)
    return img


def append_entry(path, date_str, content, photo_paths=()):
    """출하이슈사항 엑셀에 한 행 추가(A=날짜, E=내역, AU=사진). 성공 시 행 번호."""
    import openpyxl
    from openpyxl.styles import Alignment

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    dup = _find_duplicate(ws, content)
    if dup is not None:
        wb.close()
        return {"ok": True, "row": dup, "already": True}

    row = _next_row(ws)
    _ensure_row_merges(ws, row)

    try:
        d = datetime.datetime.strptime((date_str or "")[:10], "%Y-%m-%d")
    except ValueError:
        d = datetime.datetime.now()
    a = ws.cell(row=row, column=1, value=d)
    a.number_format = "yyyy-mm-dd"
    a.alignment = Alignment(horizontal="center", vertical="center")
    e = ws.cell(row=row, column=5, value=content)
    e.alignment = Alignment(vertical="center", wrap_text=True)

    n_lines = content.count("\n") + 1
    need_h = max(18 * n_lines + 6, 24)

    # 사진 — '내역' 오른쪽 AU열(47번째)에 위→아래로 쌓는다.
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    y_off = 0
    for p in photo_paths:
        try:
            img = _fit_image(p)
        except Exception:                                   # noqa: BLE001
            continue                                        # 깨진 이미지는 건너뜀
        marker = AnchorMarker(col=46, colOff=pixels_to_EMU(4),
                              row=row - 1, rowOff=pixels_to_EMU(y_off))
        img.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height)))
        ws.add_image(img)
        y_off += img.height + 8
        need_h = max(need_h, y_off * 0.75 + 6)              # px→pt 근사

    ws.row_dimensions[row].height = min(need_h, 400)

    wb.save(path)
    wb.close()
    return {"ok": True, "row": row}


def _verify_written(path, row, content):
    """저장 직후 파일을 다시 열어 그 행에 내역이 실제로 기록됐는지 확인."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        ws = wb.active
        got = None
        for r in ws.iter_rows(min_row=row, max_row=row, min_col=5, max_col=5):
            got = r[0].value
        return _norm_text(got) == _norm_text(content)
    finally:
        wb.close()


# ---------------------------------------------------------------- 잠금 사전 확인
def check_excel(model, tester_type=None):
    """이슈 저장 전 — 대상 출하이슈사항 엑셀이 열려 있는지(잠김) 미리 확인.

    저장을 눌렀는데 그제서야 실패하는 것보다, 저장 전에 '엑셀을 닫아 주세요'를
    보여주는 편이 낫다. 반환: {"ok", "locked"?, "not_found"?, "name"?, "error"?}
    """
    target = find_issue_file(model, tester_type)
    if not target.get("ok"):
        return {"ok": True, "not_found": True, "error": target.get("error")}
    if target.get("locked"):
        return {"ok": False, "locked": True, "name": target["name"],
                "error": _LOCK_MSG.format(name=target["name"])}
    return {"ok": True, "name": target["name"], "path": target["path"]}


# ---------------------------------------------------------------- 기록 행 수정
def replace_entry(path, old_content, new_content):
    """기록해 둔 행의 내역(E열)을 새 내용으로 교체. 행 못 찾으면 {"not_found"}."""
    import openpyxl
    from openpyxl.styles import Alignment
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    row = _find_duplicate(ws, old_content)
    if row is None:
        wb.close()
        return {"ok": False, "not_found": True}
    e = ws.cell(row=row, column=5, value=new_content)
    e.alignment = Alignment(vertical="center", wrap_text=True)
    n_lines = new_content.count("\n") + 1
    cur_h = ws.row_dimensions[row].height or 0
    ws.row_dimensions[row].height = min(max(18 * n_lines + 6, 24, cur_h), 400)
    wb.save(path)
    wb.close()
    return {"ok": True, "row": row}


def update_issue_text(issue_id, new_text):
    """이슈 원문(엑셀에 쓰는 그대로의 생 텍스트) 수정 — 서버 엑셀과 함께 갱신.

    · 프로그램이 기록한 행(server_export_text)이나 자동수집 원문(raw_text)과
      일치하는 엑셀 행을 찾아 내용을 교체한다.
    · 행이 없으면(수기로 지워짐 등) 새 행으로 기록한다 — 수정 저장의 의미는
      '이 내용이 서버에 있어야 한다'이므로.
    · DB 는 원문·증상/조치(재파싱)·제목·분류·태그를 함께 갱신한다.
    """
    new_text = (new_text or "").replace("\r\n", "\n").strip()
    if not new_text:
        return {"ok": False, "error": "내용이 비어 있습니다. 삭제하려면 [삭제] 버튼을 사용하세요."}

    issue = db.query("SELECT * FROM issue_history WHERE id=?", (int(issue_id),), one=True)
    if not issue:
        return {"ok": False, "error": f"이슈 #{issue_id}를 찾을 수 없습니다."}
    keys = issue.keys()
    old = (issue["server_export_text"] if "server_export_text" in keys else None) \
        or (issue["raw_text"] or "")

    target = find_issue_file(issue["model_name"], issue["tester_type"])
    if not target.get("ok"):
        return target                                    # 모델 폴더/파일 못 찾음 — 명확한 오류
    if target.get("locked"):
        return {"ok": False, "locked": True,
                "error": _LOCK_MSG.format(name=target["name"])}

    appended = False
    try:
        r = replace_entry(target["path"], old, new_text) if _norm_text(old) else \
            {"ok": False, "not_found": True}
        if r.get("not_found"):                           # 원본 행 없음 → 새 행으로 기록
            date_str = (issue["issue_date"] or "").strip() or \
                datetime.date.today().strftime("%Y-%m-%d")
            r = append_entry(target["path"], date_str, new_text)
            appended = True
    except PermissionError:
        return {"ok": False, "locked": True,
                "error": _LOCK_MSG.format(name=target["name"])}
    except Exception as e:                               # noqa: BLE001
        return {"ok": False, "error":
                f"서버 엑셀 수정 중 오류가 발생했습니다: {e}\n파일: {target['path']}"}

    try:
        if not _verify_written(target["path"], r["row"], new_text):
            return {"ok": False, "error":
                    "수정 후 검증에 실패했습니다(파일에서 수정한 내용을 찾지 못함).\n"
                    f"파일을 직접 확인해 주세요: {target['path']}"}
    except Exception:                                    # noqa: BLE001
        pass

    # DB 갱신 — 원문 그대로 + 목록 표시용 필드 재구성
    from app import tagging
    import importlib.util
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_file_location(
        "knk_import_issues", os.path.join(base, "tools", "import_issues.py"))
    imp = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(imp)
    unit, symptoms, actions = imp.split_symptom_action(new_text)
    title = imp.make_title(unit, " / ".join(symptoms), new_text)
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    db.execute(
        "UPDATE issue_history SET raw_text=?, symptom=?, action=?, title=?, item=?, "
        "unit_label=COALESCE(NULLIF(?,''), unit_label), symptom_type=?, tags=?, "
        "server_export=?, server_export_text=?, updated_at=? WHERE id=?",
        (new_text, " / ".join(symptoms) or new_text, " / ".join(actions), title, title,
         unit, tagging.classify_category(new_text),
         tagging.tags_field(tagging.auto_tags(new_text)),
         f"{target['path']} @ {now}", new_text, now, int(issue_id)))
    from app import api
    api.audit("이슈 원문수정", f"{issue['model_name']} #{issue_id}",
              f"{os.path.basename(target['path'])} {r['row']}행"
              + (" (원본 행 없음 — 새로 기록)" if appended else ""))
    return {"ok": True, "row": r["row"], "appended": appended, "path": target["path"]}


# ---------------------------------------------------------------- 기록 행 삭제
def remove_entry(path, content):
    """엑셀에서 이 프로그램이 기록한 행(내용 일치)을 비운다 — 이슈 삭제 시 사용.

    행을 지우지 않고 '비우기'만 한다(행 삭제는 아래 병합·서식을 밀어 깨뜨린다).
    빈 행은 다음 기록 때 그대로 재사용된다. 그 행에 앵커된 사진도 제거.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    row = _find_duplicate(ws, content)
    if row is None:
        wb.close()
        return {"ok": True, "not_found": True}       # 엑셀에 이미 없음 — 지울 것 없음
    ws.cell(row=row, column=1).value = None
    ws.cell(row=row, column=5).value = None
    ws._images = [im for im in ws._images
                  if getattr(getattr(im.anchor, "_from", None), "row", None) != row - 1]
    ws.row_dimensions[row].height = None
    wb.save(path)
    wb.close()
    return {"ok": True, "row": row}


def remove_exported(issue):
    """이슈 삭제 전 — 서버 엑셀에 기록해 둔 행을 함께 제거한다.

    엑셀에 행이 남은 채 프로그램에서만 지우면, 다음 서버 동기화가 그 행을
    '자동수집' 이슈로 되살린다. 그래서 엑셀 제거가 실패하면 삭제를 중단시킨다.
    """
    keys = issue.keys()
    content = issue["server_export_text"] if "server_export_text" in keys else None
    marker = issue["server_export"] if "server_export" in keys else None
    if not content:
        return {"ok": True, "not_found": True}       # 서버에 기록한 적 없음
    path = (marker or "").rsplit(" @ ", 1)[0].strip()
    if not path or not os.path.isfile(path):
        return {"ok": True, "not_found": True}       # 파일이 이동/삭제됨 — 지울 것 없음
    d, n = os.path.split(path)
    if os.path.exists(os.path.join(d, "~$" + n)):
        return {"ok": False, "locked": True, "error": _LOCK_MSG.format(name=n)}
    try:
        return remove_entry(path, content)
    except PermissionError:
        return {"ok": False, "locked": True, "error": _LOCK_MSG.format(name=n)}
    except Exception as e:                            # noqa: BLE001
        return {"ok": False, "error": f"서버 엑셀에서 행을 지우지 못했습니다: {e}\n파일: {path}"}


# ---------------------------------------------------------------- 진입점
def export_issue(issue_id):
    """이슈 1건을 서버 출하이슈사항 엑셀에 기록하고 결과를 반환.

    성공: {"ok": True, "path", "row", "already"?}
    실패: {"ok": False, "error": <사용자에게 보여줄 한국어 안내>, "locked"?/"not_found"?}
    """
    issue = db.query("SELECT * FROM issue_history WHERE id=?", (int(issue_id),), one=True)
    if not issue:
        return {"ok": False, "error": f"이슈 #{issue_id}를 찾을 수 없습니다."}
    model = (issue["model_name"] or "").strip()
    if not model:
        return {"ok": False, "not_found": True,
                "error": "이슈에 모델명이 없어 서버에 기록할 수 없습니다. 모델명을 입력해 주세요."}

    target = find_issue_file(model, issue["tester_type"])
    if not target.get("ok"):
        return target
    if target.get("locked"):
        return {"ok": False, "locked": True,
                "error": _LOCK_MSG.format(name=target["name"]), "path": target["path"]}

    content = build_content(issue)
    if not _norm_text(content):
        return {"ok": False, "error": "기록할 내용이 비어 있습니다. 증상을 입력해 주세요."}

    # 첨부 사진 → 실제 파일 경로 목록
    photos = []
    for p in db.query("SELECT file_path FROM issue_photo WHERE issue_id=? ORDER BY id",
                      (int(issue_id),)):
        fp = os.path.join(db.DATA_DIR, p["file_path"].replace("/", os.sep))
        if os.path.isfile(fp):
            photos.append(fp)

    date_str = (issue["issue_date"] or "").strip() or \
        datetime.date.today().strftime("%Y-%m-%d")

    try:
        res = append_entry(target["path"], date_str, content, photos)
    except PermissionError:
        return {"ok": False, "locked": True,
                "error": _LOCK_MSG.format(name=target["name"]), "path": target["path"]}
    except Exception as e:                                   # noqa: BLE001
        return {"ok": False, "error":
                f"서버 엑셀 기록 중 오류가 발생했습니다: {e}\n파일: {target['path']}"}

    # 기록 검증 — '저장됐다고 생각했는데 파일에 없음'을 잡는 마지막 안전망
    try:
        if not res.get("already") and not _verify_written(target["path"], res["row"], content):
            return {"ok": False, "error":
                    "기록 후 검증에 실패했습니다(파일에서 방금 쓴 내용을 찾지 못함).\n"
                    f"파일을 직접 확인해 주세요: {target['path']}"}
    except Exception:                                        # noqa: BLE001
        pass                                                 # 검증 자체 실패는 기록 성공을 뒤집지 않음

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # server_export_text: 엑셀에 실제로 기록한 내역 원문 — 이후 '서버 동기화'가
    # 같은 엑셀을 다시 읽을 때 이 내용과 일치하는 행은 건너뛰어 중복 등록을 막는다.
    db.execute("UPDATE issue_history SET server_export=?, server_export_text=? WHERE id=?",
               (f"{target['path']} @ {now}", content, int(issue_id)))
    from app import api
    api.audit("서버 이슈 기록", f"{model} #{issue_id}",
              f"{os.path.basename(target['path'])} {res['row']}행"
              + (" (기존 내용과 동일 — 중복 기록 안 함)" if res.get("already") else ""))
    return {"ok": True, "path": target["path"], "row": res["row"],
            "already": bool(res.get("already")), "photos": len(photos)}
